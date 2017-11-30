#coding=utf-8
#__author__ ="Charles.Xie"
import openpyxl
# from openpyxl.cell import get_column_letter ,column_index_from_string
wb=openpyxl.load_workbook('暂估分析测试.xlsx')
print(wb)
print(wb.get_sheet_names())
sheet=wb.get_sheet_by_name('明细')
print(sheet.columns)
# for i in range(2,17):
#     print(sheet['D'+str(i)].coordinate,sheet['D'+str(i)].value)
# sheet['d2']=1000000
# wb.save('测试.xlsx')
print(sheet.max_row,sheet.max_column)
for cust in range